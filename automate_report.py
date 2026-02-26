"""
⚡ Excel Report Automation — 3 Hours → 12 Seconds
==================================================
Transforms messy CSV/Excel data into a clean, formatted, multi-sheet 
Excel report with summary statistics, pivot tables, and charts.

Author: Deepam Shah (https://linkedin.com/in/deepammshah)
License: MIT
"""

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import sys
import time
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION — Edit these to match YOUR data
# ============================================================
CONFIG = {
    "date_columns": ["order_date", "ship_date"],
    "currency_columns": ["revenue", "cost", "unit_price"],
    "category_column": "region",
    "value_column": "revenue",
    "id_column": "order_id",
    "product_column": "product",
    "status_column": "status",
}


def print_step(step_num, description):
    """Print a formatted step indicator."""
    print(f"  [{step_num}/6] {description}")


def load_data(filepath):
    """Load CSV or Excel file into a pandas DataFrame."""
    if filepath.endswith('.csv'):
        return pd.read_csv(filepath)
    elif filepath.endswith(('.xlsx', '.xls')):
        return pd.read_excel(filepath)
    else:
        raise ValueError(f"Unsupported file format: {filepath}")


def clean_data(df):
    """
    Clean the data:
    - Remove duplicates
    - Fix date formats
    - Standardize text casing
    - Handle missing values
    - Strip whitespace
    - Remove invalid rows
    """
    original_rows = len(df)
    
    # Strip whitespace from string columns
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace('nan', pd.NA)
    
    # Remove exact duplicate rows
    df = df.drop_duplicates()
    duplicates_removed = original_rows - len(df)
    
    # Standardize text casing for categorical columns
    text_cols_to_title = ['customer_name', 'product', 'status']
    for col in text_cols_to_title:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.title()
            df[col] = df[col].replace('Nan', pd.NA)
    
    # Standardize region/category to Title Case
    if CONFIG["category_column"] in df.columns:
        df[CONFIG["category_column"]] = (
            df[CONFIG["category_column"]]
            .astype(str).str.strip().str.upper()
            .replace('NAN', pd.NA)
        )
    
    if 'category' in df.columns:
        df['category'] = df['category'].astype(str).str.strip().str.title()
        df['category'] = df['category'].replace('Nan', pd.NA)
    
    # Parse dates with multiple formats
    for date_col in CONFIG["date_columns"]:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], format='mixed', dayfirst=False, errors='coerce')
    
    # Convert currency columns to numeric
    for curr_col in CONFIG["currency_columns"]:
        if curr_col in df.columns:
            df[curr_col] = pd.to_numeric(df[curr_col], errors='coerce')
    
    # Convert quantity to numeric
    if 'quantity' in df.columns:
        df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    
    # Remove rows with negative quantities (returns) — flag them separately
    if 'quantity' in df.columns:
        returns = df[df['quantity'] < 0].copy()
        df = df[df['quantity'] >= 0]
    
    # Fill missing customer names
    if 'customer_name' in df.columns:
        df['customer_name'] = df['customer_name'].fillna('Unknown Customer')
    
    # Calculate profit margin
    if 'revenue' in df.columns and 'cost' in df.columns:
        df['profit'] = df['revenue'] - df['cost']
        df['margin_pct'] = ((df['profit'] / df['revenue']) * 100).round(1)
    
    rows_after = len(df)
    
    return df, {
        "original_rows": original_rows,
        "duplicates_removed": duplicates_removed,
        "invalid_removed": original_rows - duplicates_removed - rows_after,
        "clean_rows": rows_after,
    }


def generate_summary(df):
    """Generate summary statistics."""
    summary = {}
    
    val_col = CONFIG["value_column"]
    cat_col = CONFIG["category_column"]
    
    # Overall metrics
    summary['total_revenue'] = df[val_col].sum()
    summary['total_orders'] = len(df)
    summary['avg_order_value'] = df[val_col].mean()
    summary['total_profit'] = df['profit'].sum() if 'profit' in df.columns else 0
    summary['avg_margin'] = df['margin_pct'].mean() if 'margin_pct' in df.columns else 0
    summary['unique_customers'] = df['customer_name'].nunique()
    
    # By region
    summary['by_region'] = (
        df.groupby(cat_col)
        .agg({val_col: ['sum', 'count', 'mean']})
        .round(2)
    )
    summary['by_region'].columns = ['Total Revenue', 'Orders', 'Avg Order Value']
    summary['by_region'] = summary['by_region'].sort_values('Total Revenue', ascending=False)
    
    # By product
    if CONFIG["product_column"] in df.columns:
        summary['by_product'] = (
            df.groupby(CONFIG["product_column"])
            .agg({val_col: ['sum', 'count', 'mean']})
            .round(2)
        )
        summary['by_product'].columns = ['Total Revenue', 'Orders', 'Avg Order Value']
        summary['by_product'] = summary['by_product'].sort_values('Total Revenue', ascending=False)
    
    # Top customers
    summary['top_customers'] = (
        df.groupby('customer_name')[val_col]
        .sum()
        .sort_values(ascending=False)
        .head(5)
    )
    
    return summary


def create_charts(df, output_dir):
    """Generate charts and save as PNG files."""
    charts = []
    val_col = CONFIG["value_column"]
    cat_col = CONFIG["category_column"]
    
    plt.style.use('seaborn-v0_8-darkgrid')
    colors = ['#2563eb', '#7c3aed', '#db2777', '#ea580c', '#16a34a', '#0891b2']
    
    # Chart 1: Revenue by Region (Bar)
    fig, ax = plt.subplots(figsize=(10, 5))
    region_data = df.groupby(cat_col)[val_col].sum().sort_values(ascending=True)
    bars = ax.barh(region_data.index, region_data.values, color=colors[:len(region_data)])
    ax.set_title('Revenue by Region', fontsize=16, fontweight='bold', pad=15)
    ax.set_xlabel('Revenue (€)', fontsize=12)
    for bar, val in zip(bars, region_data.values):
        ax.text(val + 50, bar.get_y() + bar.get_height()/2, 
                f'€{val:,.0f}', va='center', fontsize=10, fontweight='bold')
    plt.tight_layout()
    chart_path = os.path.join(output_dir, 'chart_revenue_by_region.png')
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    charts.append(chart_path)
    
    # Chart 2: Revenue by Product (Pie)
    fig, ax = plt.subplots(figsize=(8, 8))
    product_data = df.groupby(CONFIG["product_column"])[val_col].sum()
    wedges, texts, autotexts = ax.pie(
        product_data.values, labels=product_data.index,
        autopct='%1.1f%%', colors=colors[:len(product_data)],
        textprops={'fontsize': 11}
    )
    for autotext in autotexts:
        autotext.set_fontweight('bold')
    ax.set_title('Revenue Share by Product', fontsize=16, fontweight='bold')
    plt.tight_layout()
    chart_path = os.path.join(output_dir, 'chart_revenue_by_product.png')
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    charts.append(chart_path)
    
    # Chart 3: Daily Revenue Trend (Line)
    if 'order_date' in df.columns:
        fig, ax = plt.subplots(figsize=(12, 5))
        daily = df.groupby(df['order_date'].dt.date)[val_col].sum()
        ax.plot(daily.index, daily.values, color=colors[0], linewidth=2.5, marker='o', markersize=6)
        ax.fill_between(daily.index, daily.values, alpha=0.15, color=colors[0])
        ax.set_title('Daily Revenue Trend', fontsize=16, fontweight='bold', pad=15)
        ax.set_ylabel('Revenue (€)', fontsize=12)
        ax.set_xlabel('Date', fontsize=12)
        plt.xticks(rotation=45)
        plt.tight_layout()
        chart_path = os.path.join(output_dir, 'chart_daily_trend.png')
        plt.savefig(chart_path, dpi=150, bbox_inches='tight')
        plt.close()
        charts.append(chart_path)
    
    # Chart 4: Top 5 Customers (Horizontal Bar)
    fig, ax = plt.subplots(figsize=(10, 5))
    top_cust = df.groupby('customer_name')[val_col].sum().sort_values(ascending=True).tail(5)
    bars = ax.barh(top_cust.index, top_cust.values, color=colors[1])
    ax.set_title('Top 5 Customers by Revenue', fontsize=16, fontweight='bold', pad=15)
    ax.set_xlabel('Revenue (€)', fontsize=12)
    for bar, val in zip(bars, top_cust.values):
        ax.text(val + 50, bar.get_y() + bar.get_height()/2,
                f'€{val:,.0f}', va='center', fontsize=10, fontweight='bold')
    plt.tight_layout()
    chart_path = os.path.join(output_dir, 'chart_top_customers.png')
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    charts.append(chart_path)
    
    return charts


def style_excel(filepath, summary, clean_stats):
    """Apply professional formatting to the Excel report."""
    wb = load_workbook(filepath)
    
    # Define styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1a1a2e')
    currency_format = '#,##0.00'
    pct_format = '0.0%'
    border = Border(
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Style each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Auto-width columns
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)
        
        # Style headers (first row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
    
    # Add Summary Dashboard sheet
    ws_dash = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws_dash['A1'] = '⚡ AUTOMATED REPORT DASHBOARD'
    ws_dash['A1'].font = Font(name='Calibri', bold=True, size=18, color='1a1a2e')
    ws_dash.merge_cells('A1:D1')
    
    ws_dash['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws_dash['A2'].font = Font(name='Calibri', size=10, color='666666')
    
    # Cleaning Stats
    ws_dash['A4'] = 'DATA CLEANING RESULTS'
    ws_dash['A4'].font = Font(name='Calibri', bold=True, size=13, color='1a1a2e')
    
    stats_data = [
        ('Original Rows', clean_stats['original_rows']),
        ('Duplicates Removed', clean_stats['duplicates_removed']),
        ('Invalid Rows Removed', clean_stats['invalid_removed']),
        ('Clean Rows', clean_stats['clean_rows']),
    ]
    
    for i, (label, value) in enumerate(stats_data, 5):
        ws_dash[f'A{i}'] = label
        ws_dash[f'A{i}'].font = Font(name='Calibri', size=11)
        ws_dash[f'B{i}'] = value
        ws_dash[f'B{i}'].font = Font(name='Calibri', bold=True, size=11)
    
    # Key Metrics
    ws_dash['A10'] = 'KEY METRICS'
    ws_dash['A10'].font = Font(name='Calibri', bold=True, size=13, color='1a1a2e')
    
    metrics = [
        ('Total Revenue', f"€{summary['total_revenue']:,.2f}"),
        ('Total Orders', summary['total_orders']),
        ('Avg Order Value', f"€{summary['avg_order_value']:,.2f}"),
        ('Total Profit', f"€{summary['total_profit']:,.2f}"),
        ('Avg Margin', f"{summary['avg_margin']:.1f}%"),
        ('Unique Customers', summary['unique_customers']),
    ]
    
    for i, (label, value) in enumerate(metrics, 11):
        ws_dash[f'A{i}'] = label
        ws_dash[f'A{i}'].font = Font(name='Calibri', size=11)
        ws_dash[f'B{i}'] = str(value)
        ws_dash[f'B{i}'].font = Font(name='Calibri', bold=True, size=11, color='2563eb')
    
    # Set column widths for dashboard
    ws_dash.column_dimensions['A'].width = 25
    ws_dash.column_dimensions['B'].width = 20
    
    wb.save(filepath)


def main():
    parser = argparse.ArgumentParser(description='⚡ Excel Report Automation — 3 Hours → 12 Seconds')
    parser.add_argument('--input', '-i', type=str, default=None,
                       help='Path to input CSV or Excel file')
    args = parser.parse_args()
    
    # Determine input file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    if args.input:
        input_file = args.input
    else:
        input_file = os.path.join(script_dir, 'sample_data', 'messy_sales_data.csv')
    
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        sys.exit(1)
    
    # Create output directory
    output_dir = os.path.join(script_dir, 'output')
    os.makedirs(output_dir, exist_ok=True)
    
    print()
    print("=" * 60)
    print("  ⚡ EXCEL REPORT AUTOMATION")
    print("  3 Hours of Manual Work → 12 Seconds")
    print("=" * 60)
    print(f"\n  Input: {os.path.basename(input_file)}")
    print()
    
    start_time = time.time()
    
    # Step 1: Load data
    print_step(1, "Loading data...")
    df = load_data(input_file)
    print(f"         → Loaded {len(df)} rows, {len(df.columns)} columns")
    
    # Step 2: Clean data
    print_step(2, "Cleaning data (duplicates, formats, missing values)...")
    df_clean, clean_stats = clean_data(df)
    print(f"         → Removed {clean_stats['duplicates_removed']} duplicates")
    print(f"         → Removed {clean_stats['invalid_removed']} invalid rows")
    print(f"         → {clean_stats['clean_rows']} clean rows remaining")
    
    # Step 3: Generate summary
    print_step(3, "Generating summary statistics...")
    summary = generate_summary(df_clean)
    print(f"         → Total Revenue: €{summary['total_revenue']:,.2f}")
    print(f"         → Avg Order Value: €{summary['avg_order_value']:,.2f}")
    
    # Step 4: Create charts
    print_step(4, "Creating charts (bar, pie, line, top customers)...")
    charts = create_charts(df_clean, output_dir)
    print(f"         → Generated {len(charts)} charts")
    
    # Step 5: Build Excel report
    print_step(5, "Building multi-sheet Excel report...")
    report_name = f"report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    report_path = os.path.join(output_dir, report_name)
    
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        # Sheet 1: Clean Data
        df_clean.to_excel(writer, sheet_name='Clean Data', index=False)
        
        # Sheet 2: By Region
        summary['by_region'].to_excel(writer, sheet_name='By Region')
        
        # Sheet 3: By Product
        if 'by_product' in summary:
            summary['by_product'].to_excel(writer, sheet_name='By Product')
        
        # Sheet 4: Top Customers
        summary['top_customers'].to_frame('Total Revenue').to_excel(
            writer, sheet_name='Top Customers'
        )
    
    # Step 6: Apply formatting
    print_step(6, "Applying professional formatting & dashboard...")
    style_excel(report_path, summary, clean_stats)
    
    elapsed = time.time() - start_time
    
    print()
    print("=" * 60)
    print(f"  ✅ DONE in {elapsed:.1f} seconds!")
    print("=" * 60)
    print(f"\n  📊 Report: {report_path}")
    print(f"  📈 Charts: {output_dir}/chart_*.png")
    print(f"\n  What would have taken ~3 hours manually")
    print(f"  was completed in {elapsed:.1f} seconds.")
    print()


if __name__ == "__main__":
    main()
